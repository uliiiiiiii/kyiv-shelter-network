import requests
import psycopg2
from openpyxl import load_workbook
import os
from dotenv import load_dotenv

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")
GEOCODING_API_KEY = os.getenv("GEOCODING_API_KEY")

def geocode_address(address):
    url = f"https://api.mapbox.com/geocoding/v5/mapbox.places/{address}.json?access_token={GEOCODING_API_KEY}"
    response = requests.get(url)
    json = response.json()
    
    if json['features']:
        # Mapbox returns longitude,latitude; so reverse the order
        longitude, latitude = json['features'][0]['geometry']['coordinates']
        return latitude, longitude
    else:
        raise Exception(f"Could not geocode address: {address}")


def connect_to_postgres():
    return psycopg2.connect(DATABASE_URL)

def process_and_insert_data():
    wb = load_workbook('shelters.xlsx')
    sheet = wb.active
    
    conn = connect_to_postgres()
    cur = conn.cursor()

    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            district, address, shelter_type, place, building_type, owner, ownership, phone, accessibility, hours, *_  = row
            
            accessibility = accessibility != "Відсутній"
            
            try:
                latitude, longitude = geocode_address(f"{address}, Київ")
            except Exception as e:
                print(f"Error geocoding {address}: {e}")
                continue 
           
            query = """
                INSERT INTO "Shelter" (district, address, shelter_type, place, building_type, owner, ownership, phone, accessibility, hours, latitude, longitude)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            data = (
                district,
                address,
                shelter_type,
                place,
                building_type,
                owner,
                ownership,
                phone,
                accessibility,
                hours,
                latitude,
                longitude
            )
            
            cur.execute(query, data)
        
        conn.commit()
        print("Data successfully inserted into the database.")
    except Exception as e:
        print(f"Error processing data: {e}")
        conn.rollback() 
    finally:
        cur.close()
        conn.close()

process_and_insert_data()
